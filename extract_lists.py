####################################################
#                  Revision: 1.1                   #
# What's new: Changed naming constraints on        # 
# Supported drives file. Now it can be             # 
# "Supported_Drives", "Supported_drives",          #
# "supported_Drives", or "supported_drives".       #
#                                                  #
#              Updated on: 11/06/2015              #
####################################################


####################################################
#                  Revision: 1.0                   #
#              Updated on: 11/06/2015              #
####################################################

####################################################
#                                                  #
#   This File extracts lists of Vendors, OEM       #
#   models,Vendor Internal Names, Vendor Family    #
#   names, Capacities, Next FW revs, and Release   #
#   ECOs from "Supported_Drive.xlsx" located       #
#   in Local directory. This Model List is used    #
#   to validate with User Inputs. And other data   # 
#   lists are used in Report.                      #
#                                                  #
#   Author: Zankar Sanghavi                        #
#                                                  #
#   © Dot Hill Systems Corporation                 #
#                                                  #
####################################################


###################################################
#
#   Importing required packages
# 
###################################################

from openpyxl import load_workbook
import report_functions
import fixed_data
import warnings
warnings.filterwarnings("ignore")
class Extract_Lists:
    def load_in_wb(directory, string):
        wb = load_workbook(filename = r''+str(directory)
                            +str(string))
        return wb                    

    def get_data():
        ###################################################
        #
        #   Loading Supported_drives workbook and extracting 
        #   required columns from '3.5 BB', '2.5 SSD', 
        #   '2.5 BB', and '2.5 HP' worksheets.
        #
        #   We are not using 'Teradata' worksheet. As it 
        #   contains data reference from other 4 worksheets
        #   mentioned above. If in Future, we want to add 
        #   this worksheet, one should make sure that it 
        #   should have typed data.
        #    
        ###################################################

        ###############################
        #
        #   Empty Lists to store data
        #
        ###############################
        model_data=[] #Model data for BB & SSD
        model_data_HP=[] #Model data for BB & SSD

        capacity_data=[] # BB & SSD
        capacity_data_HP=[] # For HP

        vendor_data=[]# BB & SSD
        vendor_data_HP=[]# For HP

        product_fname_data=[]# BB & SSD
        product_fname_data_HP=[]# For HP

        fw_data=[]# BB & SSD
        fw_data_HP=[]# For HP

        eco_data=[]
        eco_data_HP=[]# For HP

        fd = fixed_data.Fixed_Data
        
        #wb = Extract_Lists.load_in_wb(directory, string)
        # wb = load_workbook(filename = r''+str(fd.fixed_dir)
                            # +'\Supported_Drives.xlsx')

        possible_file_name = ['\Supported_Drives.xlsx', '\Supported_drives.xlsx', '\supported_Drives.xlsx', '\supported_drives.xlsx']
        
        for i in range(len(possible_file_name)):
            try:
                wb = Extract_Lists.load_in_wb(fd.fixed_dir, possible_file_name[i])
                break
            except:
                continue
                
        ws_1 = wb['3.5 BB'] # Names of Worksheet
        ws_2 = wb['2.5 SSD']
        ws_3 = wb['2.5 BB']
        ws_4 = wb['2.5 HP']
        #ws_5 = wb['Teradata']

        
        ###################################################
        #
        #   Extracting required columns as mentioned above
        # 
        ###################################################
        ed= report_functions.Report_Functions
        
        #For BB & SSD
        non_hp_ws=[ws_1,ws_2,ws_3]

        for i in range(len(non_hp_ws)):
            model_data=ed.extract_data(non_hp_ws[i]
                                        ,'E',model_data)
                                        
            capacity_data=ed.extract_data(non_hp_ws[i]
                                            ,'I',capacity_data)
                                            
            vendor_data=ed.extract_data(non_hp_ws[i]
                                        ,'B',vendor_data)
                                        
            product_fname_data= ed.extract_data(non_hp_ws[i]
                                        ,'G',product_fname_data)
                                        
            fw_data=ed.extract_data(non_hp_ws[i],'M',fw_data)
            eco_data=ed.extract_data(non_hp_ws[i],'Q',eco_data)

        model_list=model_data
        capacity_list=capacity_data
        vendor_list=vendor_data
        product_fname_list=product_fname_data
        fw_list=fw_data
        eco_list=eco_data

        #For HP    
        model_data_HP=ed.extract_data(ws_4,'D',model_data_HP) 
        capacity_data_HP=ed.extract_data(ws_4,'I'
                                        ,capacity_data_HP)
        vendor_data_HP=ed.extract_data(ws_4,'B'
                                        ,vendor_data_HP)
        product_fname_data_HP= ed.extract_data(ws_4,'G'
                                        ,product_fname_data_HP)
        fw_list_HP=ed.extract_data(ws_4,'M'
                                    ,fw_data_HP)

        eco_data_HP=ed.extract_data(ws_4,'Q',eco_data_HP)


        ########################################################
        #
        #   openpyxl module reads extra 100-150 vacant 
        #   lines approximately in 4th worksheet (i.e. 2.5 HP).
        #   So in order to eliminate those None values.
        #   Here is a little loop, which takes care of it.   
        # 
        #########################################################

        index=[]
        for i in range(len(model_data_HP)):
            if model_data_HP[i]==None:
                #print(i)
                index.append(i)


        model_list_HP=model_data_HP[:min(index)]   
        #final model list from Excel file

        capacity_list_HP=capacity_data_HP[:min(index)] 
        #final capacity list from Excel file

        vendor_list_HP=vendor_data_HP[:min(index)] 
        #final Vendor list from Excel file

        fw_list_HP=fw_data_HP[:min(index)] 
        #final Firmware list from Excel file

        eco_list_HP=eco_data_HP[:min(index)] 
        #final Release ECO list from Excel file

        product_fname_list_HP=product_fname_data_HP[:min(index)] #final Product name list from Excel file
        return [model_list, capacity_list, vendor_list, fw_list
                , eco_list, product_fname_list,model_list_HP
                , capacity_list_HP, vendor_list_HP, fw_list_HP
                , eco_list_HP, product_fname_list_HP]
            
            
#####################################
#              END                  #
##################################### 